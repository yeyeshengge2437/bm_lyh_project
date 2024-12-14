import base64
import os
import random

import requests
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl import Workbook

def get_excel_data(url):
    num_name = random.randint(12345, 99999)
    file_type = url.split('.')[-1]
    if file_type == 'xls':
        file_name_old = f"{num_name}.xls"
        file_name_new = f"{num_name}_cleaned.xls"
        file = requests.get(url)
        with open(file_name_old, "bw") as f:
            f.write(file.content)

        # 打开工作簿
        rb = xlrd.open_workbook(file_name_old, formatting_info=False)


        # 复制工作簿以避免直接修改原始文件
        wb = copy(rb)

        # 遍历所有工作表
        for sheet_index in range(rb.nsheets):
            ws = rb.sheet_by_index(sheet_index)
            w_ws = wb.get_sheet(sheet_index)

            # 遍历所有行
            for row_index in range(ws.nrows):
                # 遍历所有列
                for col_index in range(ws.ncols):
                    cell = ws.cell(row_index, col_index)
                    # 获取单元格的值
                    value = cell.value
                    # 直接写入值，不保留任何格式
                    w_ws.write(row_index, col_index, value)

        # 保存处理后的工作簿
        wb.save(file_name_new)

        with open(file_name_new, "br") as f:
            excel_value = f.read()
        excel_value = base64.b64encode(excel_value).decode()
        # 删除临时文件
        os.remove(file_name_old)
        os.remove(file_name_new)
        return excel_value
    elif file_type == 'xlsx':
        file_name_old = f"{num_name}.xlsx"
        file_name_new = f"{num_name}_cleaned.xlsx"
        file = requests.get(url)
        with open(file_name_old, "bw") as f:
            f.write(file.content)

        # 加载现有的工作簿
        wb = load_workbook(filename=file_name_old)

        # 创建一个新的工作簿
        new_wb = Workbook()

        # 将新工作簿的当前工作表重命名（如果需要）
        new_wb.active.title = wb.active.title

        # 遍历所有工作表
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            new_ws = new_wb.create_sheet(sheet)

            # 遍历所有行
            for row in ws.iter_rows():
                # 创建一行新单元格
                new_row = []
                for cell in row:
                    # 获取单元格的值
                    new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    new_row.append(new_cell)
                new_ws.append(new_row)

        # 保存新工作簿
        new_wb.save(filename=file_name_new)

        with open(file_name_new, "br") as f:
            excel_value = f.read()
        excel_value = base64.b64encode(excel_value).decode()
        # 删除临时文件
        os.remove(file_name_old)
        os.remove(file_name_new)
        return excel_value
    else:
        return False


# print(get_excel_value("https://res.debtop.com/col/live/paper/202410/24/2024102419443821b0c98627e64e1b.xls"))




