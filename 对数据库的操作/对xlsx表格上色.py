import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# 计算字符串中的汉字数量
def count_chinese_chars(text):
    if not text:
        return 0
    return sum(1 for char in text if '\u4e00' <= char <= '\u9fff')


# 加载Excel文件并设置背景色
def add_background_color_by_row_col(file_path, start_row, end_row, start_col, end_col, color):
    # 加载工作簿
    wb = load_workbook(file_path)
    # 选择工作表
    ws = wb.active  # 或者使用 wb.get_sheet_by_name('Sheet1') 如果不是默认工作表

    # 设置填充样式
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    # 应用填充样式到指定行和列的单元格
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill

    # 保存工作簿
    wb.save(file_path)


# add_background_color_by_row_col("eeee.xlsx", 1, 1, 1, 1, "FFFF0000")


# 比较两列的汉字数量，并填充红色背景
def compare_columns_and_fill(file_path, sheet_name=0, col1=None, col2=None, color="FFFF0000"):
    df = pd.read_excel(file_path, usecols=f"{col1}, {col2}", )
    for index, row in df.iterrows():
        guarantee, guarantee_new = row
        if str(guarantee) == 'nan':
            guarantee = ""
        if str(guarantee) == '无':
            guarantee = ""
        if str(guarantee_new) == '空':
            guarantee_new = ""
        print(guarantee, 111)
        guarantee_num = count_chinese_chars(guarantee)
        print(guarantee_new, 222)
        guarantee_new_num = count_chinese_chars(guarantee_new)
        if guarantee_num == guarantee_new_num:
            print(index)
        elif guarantee_num > guarantee_new_num:
            add_background_color_by_row_col(file_path, int(index) + 2, int(index) + 2, 10, 10, color)  # 红色背景
            print("guarantee_new多")
        else:
            add_background_color_by_row_col(file_path, int(index) + 2, int(index) + 2, 9, 9, color)  # 红色背景
            print("guarantee_new少")


# color="00EE90" # 淡绿色
# color="FFFF0000" # 红色
# color="FFFF00"

compare_columns_and_fill("deepseek.xlsx", sheet_name=0, col1="I", col2="J", color='00EE90')
