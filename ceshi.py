from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO

# 加载Excel工作簿
wb = load_workbook('1111.xlsx')
sheet = wb.active

# 将Excel工作表转换为图片
img = Image.new('RGB', (sheet.max_column, sheet.max_row), color='white')
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        img.paste(OpenpyxlImage(cell.value), (cell.column-1, cell.row-1))

# 保存图片
img.save('excel_image.png')