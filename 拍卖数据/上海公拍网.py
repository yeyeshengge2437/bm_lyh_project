import re
import time

from openpyxl import Workbook
import requests
from lxml import etree

# 创建一个Workbook对象，它代表一个Excel文件
wb = Workbook()

# 选择默认的工作表
ws = wb.active

# 给工作表命名
ws.title = "上海公拍网"

# 写入数据到单元格
ws['A1'] = '标题'
ws['B1'] = '链接'
ws['C1'] = '起拍价(元)'
ws['D1'] = '评估价(元)'
ws['E1'] = '开始时间'
ws['F1'] = '结束时间'
ws['G1'] = '预计结束时间'
ws['H1'] = '状态'
ws['I1'] = '所在地区'
ws['J1'] = '标的类型'
xlsx_data = []

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    # 'Cookie': 'Hm_lvt_263a15f1b2e57ebc22960d3fa7c5537e=1730077707; HMACCOUNT=FDD970C8B3C27398; ASPSESSIONIDAADTTRAQ=FPECAKGDCMKFJNBOPEEHHJBE; SF_cookie_67=18454395; dLwyZZHe134zS=60meDGmkLZnngmeIgHC31_S6EgC2MwM0ER9R.MSEePl6bbUJw5oWhrqZz6tVn3SQzZ9QaPbIw8TwY7fE8Wcx7Z8A; SF_cookie_8=27902555; HMF_CI=44bb12c2e90066964c919688ad52681efa202a9be20bf27acaa80d15d25f6132921721e8dcb68d7fd4b8977fcc91461ed789fe290c29d8fdaf2b4c2b9d372b5a6e; FECW=129062a52080fa364b664b2ce8aaadb5facd226ae9b41b405a519c17eb14e2fca831bd1c17c032322362715b657f4803a5f21c70acb66608539caaacc28cb0205d17e90187867435f828a63af51c23c924; HMY_JC=53caad8322c0473b5438d616cda2cb32c908ac6a4c6a6c47a8f4960451f8880f36,; HBB_HC=2a9e762744a58a1681b967af8a10b2dcae5872f23ab37961c85ec01ff8f10e661bfaa0aa26c340f60f6075cbbfc1a44690; Hm_lpvt_263a15f1b2e57ebc22960d3fa7c5537e=1730078170; dLwyZZHe134zT=0GpirNb.a6PzyJmZ2qkEjx_sgDCVhhdC4tXOYHEH0OMH2A3DuGHB0fs2Sds_wgoG8fEy6Bf03vzl40X4T4Qsi5LyAuvKhtP1.ThgNo6ddIvlLiastRwuPi4nouXd4OMrj99owxQYXNCOSJTaboGWD.lwJIjPwzzERFFRPUNxwb5Z4goV16IxEmYmx.qdUKloudxz_TOL07YuZu9Vj6nX.QIg1wWWZneApdxXeb2_uHJawc_4pWJ1Vu_KLTOmbOEe5XVFQlXIuqaapbtoJwsjrpR2OfZghmtuBbHc2xQgAv3PS.bQ0N.O.B8Jm0efHyA.KwC5RwqFSY8Gd4fAUfbVYdvN4nQXxen93KrZ5njml8gZ; FECA=eqsYOEg4Gk4cmla6h4pQkUQcwFWtxl/k9JrZmR8241PxQYKvt8bZKeLq9O2cuBUJpkFtjNE4+8YBAP3ar3wRwNwrtKkhuWuZ4WngrTooK+rhtTJsuAh+sbnAt2CBr+3jGDv2/hnPgmGC+eCjXJdYYgAgdd7EBa8B/ihCe32zgdjBxRTsJjY5Y8GIM/NSRRXJob; C3VK=f97621',
    'Pragma': 'no-cache',
    'Referer': 'https://s.gpai.net/sf/search.do?q=&pr=447&restate=1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
}

target_type = {
    '376': '房产',
    '381': '土地',
    '372': '股权',
    '378': '机动车',
    '383': '船舶',
    '377': '物资',
    '380': '工艺品',
    '385': '无形资产',
    '382': '财产性权益',
    '386': '林权矿权',
    '379': '其他'
}
location = {
    '31': '上海',
    '3201': '南京',
    '3301': '杭州',
    '3205': '苏州',
}
for target_type_key, target_type_value in target_type.items():
    for location_key, location_value in location.items():
        page = 0
        while True:
            page += 1
            params = {
                'q': '',
                'pr': '447',
                'restate': '4',  # 拍卖状态
                'at': f'{target_type_key}',  # 标的类型
                'cityNum': f'{location_key}',  # 地区
                'Page': f'{page}',  # 页数
            }

            response = requests.get('https://s.gpai.net/sf/search.do', params=params, headers=headers)
            auction_data = response.content.decode()
            html = etree.HTML(auction_data)
            # 获取所有数据
            overall_data = html.xpath("//div[@class='block-wrap notice-wrap'][1]//ul[@class='main-col-list clearfix']/li/div[@class='list-item']")

            for data in overall_data:
                title = "".join(data.xpath(".//div[@class='item-tit']/a/text()"))
                title_url = "https:" + "".join(data.xpath(".//div[@class='item-tit']/a/@href"))
                starting_bid = "".join(data.xpath(".//p[2]//text()")).strip()
                appraisal_price = "".join(data.xpath(".//p[3]//text()")).strip()
                time_date = "".join(data.xpath(".//p[4]//text()")).strip()
                auction_status = "".join(data.xpath(".//span[contains(@class, 'badge-icon')]/text()"))
                if '预计结束' in time_date:
                    expected_end = time_date.split('预计结束：')[1]
                elif '开始时间' in time_date:
                    start_time = time_date.split('开始时间：')[1]
                elif '结束时间' in time_date:
                    end_time = time_date.split('结束时间：')[1]
                if 'expected_end' not in globals():
                    expected_end = ''
                if 'start_time' not in globals():
                    start_time = ''
                if 'end_time' not in globals():
                    end_time = ''
                if '万元' in appraisal_price:
                    appraisal_price_num = re.findall(r'评估价：(.*?)万元', appraisal_price)[0]
                    appraisal_price_num = float(appraisal_price_num) * 10000
                else:
                    appraisal_price_num = re.findall(r'评估价：(.*?)元', appraisal_price)[0]
                if '万元' in starting_bid:
                    starting_bid_num = re.findall(r'起拍价：(.*?)万元', starting_bid)[0]
                    starting_bid_num = float(starting_bid_num) * 10000
                else:

                    try:
                        starting_bid_num = re.findall(r'[\d.]+', starting_bid)[0]
                    except:
                        print(starting_bid)
                        break
                if float(starting_bid_num) >= 300000000:
                    print(title, title_url, starting_bid, appraisal_price, time_date, auction_status)
                    xlsx_data.append([title, title_url, starting_bid_num, appraisal_price_num, start_time, end_time, expected_end, auction_status, location_value, target_type_value])
            if len(overall_data) < 20:
                break
            time.sleep(2)

# 使用行和列的迭代器来写入数据
for row_idx, row_data in enumerate(xlsx_data, start=2):  # start=2表示从第二行开始写入
    for col_idx, value in enumerate(row_data, start=1):  # start=1表示从第一列开始写入
        ws.cell(row=row_idx, column=col_idx, value=value)

# 保存Workbook对象到文件
wb.save("上海公拍网2.xlsx")




